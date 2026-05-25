from __future__ import annotations

import csv
from pathlib import Path
from urllib.parse import quote_plus

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError as exc:  # pragma: no cover - local environment guard
    raise SystemExit(
        "openpyxl is required to generate the workbook. "
        "Install it or run with the bundled Codex Python environment."
    ) from exc


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"

COMPANIES = [
    {
        "company": "Salesforce",
        "ticker": "CRM",
        "industry": "Enterprise SaaS",
        "segment": "CRM / enterprise platform",
        "selected": "Yes",
        "priority": "Core",
        "data_score": 5,
        "story_score": 4,
        "peer_score": 5,
        "pilot_role": "Deterioration/improvement candidate TBD after baseline",
        "why": "Large employee base, long public history, frequent investor communication, clear SaaS peer set.",
        "ir_url": "https://investor.salesforce.com",
    },
    {
        "company": "ServiceNow",
        "ticker": "NOW",
        "industry": "Enterprise SaaS",
        "segment": "Workflow automation",
        "selected": "Yes",
        "priority": "Core",
        "data_score": 5,
        "story_score": 4,
        "peer_score": 5,
        "pilot_role": "Stable candidate TBD after baseline",
        "why": "High-quality SaaS comparator with deep employee and transcript footprint.",
        "ir_url": "https://investors.servicenow.com",
    },
    {
        "company": "Workday",
        "ticker": "WDAY",
        "industry": "Enterprise SaaS",
        "segment": "HCM / finance SaaS",
        "selected": "Yes",
        "priority": "Core",
        "data_score": 5,
        "story_score": 3,
        "peer_score": 5,
        "pilot_role": "",
        "why": "Strong HR-software peer, talent-heavy organization, useful culture baseline comparator.",
        "ir_url": "https://investor.workday.com",
    },
    {
        "company": "Snowflake",
        "ticker": "SNOW",
        "industry": "Enterprise SaaS",
        "segment": "Data cloud",
        "selected": "Yes",
        "priority": "Core",
        "data_score": 4,
        "story_score": 5,
        "peer_score": 5,
        "pilot_role": "Improvement/deterioration candidate TBD after baseline",
        "why": "Public growth-stage SaaS name with leadership/product narrative shifts and strong transcript signal.",
        "ir_url": "https://investors.snowflake.com",
    },
    {
        "company": "Datadog",
        "ticker": "DDOG",
        "industry": "Enterprise SaaS",
        "segment": "Observability / cloud monitoring",
        "selected": "Yes",
        "priority": "Core",
        "data_score": 4,
        "story_score": 4,
        "peer_score": 5,
        "pilot_role": "",
        "why": "Developer-heavy SaaS with useful talent-flow and product-language signals.",
        "ir_url": "https://investors.datadoghq.com",
    },
    {
        "company": "HubSpot",
        "ticker": "HUBS",
        "industry": "Enterprise SaaS",
        "segment": "SMB CRM / marketing automation",
        "selected": "Yes",
        "priority": "Core",
        "data_score": 4,
        "story_score": 4,
        "peer_score": 5,
        "pilot_role": "",
        "why": "Culture-sensitive SaaS company with strong employee-review signal and public investor materials.",
        "ir_url": "https://ir.hubspot.com",
    },
    {
        "company": "Atlassian",
        "ticker": "TEAM",
        "industry": "Enterprise SaaS",
        "segment": "Collaboration / developer tools",
        "selected": "Yes",
        "priority": "Core",
        "data_score": 4,
        "story_score": 4,
        "peer_score": 5,
        "pilot_role": "",
        "why": "Developer-tool peer with distributed workforce history and rich public commentary.",
        "ir_url": "https://investors.atlassian.com",
    },
    {
        "company": "MongoDB",
        "ticker": "MDB",
        "industry": "Enterprise SaaS",
        "segment": "Database / developer platform",
        "selected": "Yes",
        "priority": "Core",
        "data_score": 4,
        "story_score": 4,
        "peer_score": 5,
        "pilot_role": "",
        "why": "Developer-platform SaaS with visible enterprise transition and talent-market signal.",
        "ir_url": "https://investors.mongodb.com",
    },
    {
        "company": "NVIDIA",
        "ticker": "NVDA",
        "industry": "Semiconductor",
        "segment": "Accelerated computing / GPUs",
        "selected": "Yes",
        "priority": "Core",
        "data_score": 5,
        "story_score": 5,
        "peer_score": 5,
        "pilot_role": "Stable/improvement candidate TBD after baseline",
        "why": "Talent magnet with unusually strong public narrative; essential semiconductor reference point.",
        "ir_url": "https://investor.nvidia.com",
    },
    {
        "company": "AMD",
        "ticker": "AMD",
        "industry": "Semiconductor",
        "segment": "CPUs / GPUs / data center chips",
        "selected": "Yes",
        "priority": "Core",
        "data_score": 5,
        "story_score": 5,
        "peer_score": 5,
        "pilot_role": "",
        "why": "Direct semiconductor comparator with leadership, execution, and talent-flow relevance.",
        "ir_url": "https://ir.amd.com",
    },
    {
        "company": "Intel",
        "ticker": "INTC",
        "industry": "Semiconductor",
        "segment": "Integrated device manufacturer / foundry",
        "selected": "Yes",
        "priority": "Core",
        "data_score": 5,
        "story_score": 5,
        "peer_score": 5,
        "pilot_role": "Deterioration/improvement candidate TBD after baseline",
        "why": "Large workforce, strategic restructuring, and heavy transcript/news signal.",
        "ir_url": "https://www.intc.com",
    },
    {
        "company": "Micron",
        "ticker": "MU",
        "industry": "Semiconductor",
        "segment": "Memory",
        "selected": "Yes",
        "priority": "Core",
        "data_score": 4,
        "story_score": 4,
        "peer_score": 5,
        "pilot_role": "",
        "why": "Memory-cycle comparator with enough employee, transcript, and macro-sensitivity context.",
        "ir_url": "https://investors.micron.com",
    },
    {
        "company": "Qualcomm",
        "ticker": "QCOM",
        "industry": "Semiconductor",
        "segment": "Wireless / edge AI chips",
        "selected": "Yes",
        "priority": "Core",
        "data_score": 5,
        "story_score": 4,
        "peer_score": 5,
        "pilot_role": "",
        "why": "Large engineering organization with mobile, automotive, and AI narrative shifts.",
        "ir_url": "https://investor.qualcomm.com",
    },
    {
        "company": "Applied Materials",
        "ticker": "AMAT",
        "industry": "Semiconductor",
        "segment": "Semiconductor equipment",
        "selected": "Yes",
        "priority": "Core",
        "data_score": 4,
        "story_score": 4,
        "peer_score": 5,
        "pilot_role": "",
        "why": "Semicap equipment anchor with cyclical and geopolitical context.",
        "ir_url": "https://ir.appliedmaterials.com",
    },
    {
        "company": "Lam Research",
        "ticker": "LRCX",
        "industry": "Semiconductor",
        "segment": "Semiconductor equipment",
        "selected": "Yes",
        "priority": "Core",
        "data_score": 4,
        "story_score": 4,
        "peer_score": 5,
        "pilot_role": "",
        "why": "Semicap equipment peer for AMAT/KLA comparisons and talent-flow checks.",
        "ir_url": "https://investor.lamresearch.com",
    },
    {
        "company": "Okta",
        "ticker": "OKTA",
        "industry": "Enterprise SaaS",
        "segment": "Identity / security SaaS",
        "selected": "No",
        "priority": "Alternate",
        "data_score": 4,
        "story_score": 4,
        "peer_score": 4,
        "pilot_role": "Alternate",
        "why": "Useful backup if a core SaaS name has weak manual data coverage.",
        "ir_url": "https://investor.okta.com",
    },
    {
        "company": "Twilio",
        "ticker": "TWLO",
        "industry": "Enterprise SaaS",
        "segment": "Communications platform",
        "selected": "No",
        "priority": "Alternate",
        "data_score": 4,
        "story_score": 5,
        "peer_score": 4,
        "pilot_role": "Alternate",
        "why": "Backup with restructuring and culture-change potential.",
        "ir_url": "https://investors.twilio.com",
    },
    {
        "company": "KLA",
        "ticker": "KLAC",
        "industry": "Semiconductor",
        "segment": "Process control / metrology",
        "selected": "No",
        "priority": "Alternate",
        "data_score": 4,
        "story_score": 3,
        "peer_score": 5,
        "pilot_role": "Alternate",
        "why": "Backup semicap comparator for AMAT/Lam Research.",
        "ir_url": "https://ir.kla.com",
    },
    {
        "company": "Tesla",
        "ticker": "TSLA",
        "industry": "Automotive / EV",
        "segment": "Electric vehicles / energy / autonomy",
        "selected": "No",
        "priority": "Alternate",
        "data_score": 5,
        "story_score": 5,
        "peer_score": 3,
        "pilot_role": "Alternate (cross-sector)",
        "why": "Cross-sector alternate. Huge public employee footprint, extreme management-language volatility, frequent talent inflows/outflows; no clean SaaS or semi peer, so kept as alternate rather than core.",
        "ir_url": "https://ir.tesla.com",
    },
]

COMPANY_HEADERS = [
    "Company",
    "Ticker",
    "Industry",
    "Segment",
    "Selected",
    "Priority",
    "Data Thickness Score",
    "Story Volatility Score",
    "Peer Comparability Score",
    "Total Priority Score",
    "Pilot Role",
    "Why Included",
    "Research Status",
    "Owner",
    "Next Step",
    "Notes",
]

TRACKER_HEADERS = [
    "Judgment ID",
    "Publish Date",
    "Company",
    "Ticker",
    "Direction",
    "Confidence",
    "Difficulty",
    "Primary Metric",
    "Threshold",
    "Deadline",
    "Status",
    "Resolution Date",
    "Outcome Notes",
    "Source Links",
]

SOURCE_HEADERS = [
    "Company",
    "Ticker",
    "Source Type",
    "URL",
    "Access Notes",
    "Last Checked",
    "Evidence Line",
    "Quality Notes",
]

ALERT_HEADERS = ["Company", "Ticker", "Category", "Google Alerts Query", "Alert Frequency", "Owner", "Status"]

WEEKLY_HEADERS = ["Week", "Phase", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday Deliverable", "Acceptance Check"]


def selected_companies() -> list[dict[str, object]]:
    return [company for company in COMPANIES if company["selected"] == "Yes"]


def write_csv(path: Path, headers: list[str], rows: list[list[object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.writer(file)
        writer.writerow(headers)
        writer.writerows(rows)


def company_rows() -> list[list[object]]:
    rows = []
    for item in COMPANIES:
        total = item["data_score"] + item["story_score"] + item["peer_score"]
        rows.append(
            [
                item["company"],
                item["ticker"],
                item["industry"],
                item["segment"],
                item["selected"],
                item["priority"],
                item["data_score"],
                item["story_score"],
                item["peer_score"],
                total,
                item["pilot_role"],
                item["why"],
                "Not started",
                "",
                "Build baseline evidence log",
                "Operational prioritization only; not a culture quality conclusion.",
            ]
        )
    return rows


def source_rows() -> list[list[object]]:
    rows = []
    for item in selected_companies():
        company = str(item["company"])
        ticker = str(item["ticker"])
        encoded_company = quote_plus(company)
        rows.extend(
            [
                [
                    company,
                    ticker,
                    "Investor Relations",
                    item["ir_url"],
                    "Use for earnings releases, presentations, and official transcript links where available.",
                    "",
                    "Management Language",
                    "Primary source preferred.",
                ],
                [
                    company,
                    ticker,
                    "SEC EDGAR Search",
                    f"https://www.sec.gov/edgar/search/#/q={ticker}",
                    "Use for 10-K, 10-Q, 8-K, risk factors, restructuring disclosures.",
                    "",
                    "External Red Flags",
                    "Primary regulatory source.",
                ],
                [
                    company,
                    ticker,
                    "Glassdoor Manual Search",
                    f"https://www.google.com/search?q={encoded_company}+Glassdoor+reviews",
                    "Manual review only; record ratings and review themes, do not scrape.",
                    "",
                    "Employee Review Trend",
                    "Check overall, CEO approval, business outlook, management confidence where visible.",
                ],
                [
                    company,
                    ticker,
                    "LinkedIn People Search",
                    f"https://www.linkedin.com/search/results/people/?keywords={quote_plus(company + ' VP product engineering staff principal')}",
                    "Manual logged-in research; record VP+/Staff+/core product and engineering moves.",
                    "",
                    "Talent Flow",
                    "Most labor-intensive source; strongest differentiator.",
                ],
                [
                    company,
                    ticker,
                    "Transcript Search",
                    f"https://www.google.com/search?q={encoded_company}+earnings+call+transcript",
                    "Prefer IR, Seeking Alpha, Motley Fool, or other public transcript source.",
                    "",
                    "Management Language",
                    "Read latest 2-3 calls before publishing a card.",
                ],
                [
                    company,
                    ticker,
                    "News / Red Flag Search",
                    f"https://www.google.com/search?q={encoded_company}+layoffs+lawsuit+SEC+investigation+executive+departure",
                    "Use to reconcile Google Alerts and external events.",
                    "",
                    "External Red Flags",
                    "Record material events only.",
                ],
            ]
        )
    return rows


def alert_rows() -> list[list[object]]:
    categories = [
        ("Layoffs", "{company} layoffs"),
        ("Lawsuit", "{company} lawsuit"),
        ("SEC", "{company} SEC"),
        ("Investigation", "{company} investigation"),
        ("Culture", "{company} culture"),
        ("Executive Departure", "{company} executive departure"),
    ]
    rows = []
    for item in selected_companies():
        for category, query_template in categories:
            rows.append(
                [
                    item["company"],
                    item["ticker"],
                    category,
                    query_template.format(company=item["company"]),
                    "At most once a day",
                    "",
                    "Not configured",
                ]
            )
    return rows


def weekly_rows() -> list[list[object]]:
    return [
        [
            "0",
            "Setup",
            "Create workbook and source library.",
            "Confirm first 15 companies and 3 alternates.",
            "Set Google Alerts query list.",
            "Assign 3 pilot companies.",
            "Research base ready.",
            "Workbook exists; selected company count is 15; no judgment rows without validation metrics.",
        ],
        [
            "1",
            "Pilot research",
            "Read alerts and IR pages for pilot companies.",
            "Record Glassdoor ratings and review themes.",
            "Map LinkedIn talent-flow sample.",
            "Read latest 2-3 transcripts.",
            "3 internal draft cards.",
            "Each draft card has four evidence lines and source links.",
        ],
        [
            "2",
            "Pilot close",
            "Patch missing sources.",
            "Normalize evidence notes.",
            "Pick the strongest publishable card.",
            "Prepare public about page.",
            "Public launch package ready.",
            "Pilot workload is measured; publishing cadence is set.",
        ],
        [
            "3",
            "Launch",
            "Publish about page.",
            "Publish methodology article.",
            "Publish first card.",
            "Post bilingual X thread.",
            "Update judgment tracker.",
            "Every public claim has a timestamp and source link.",
        ],
        [
            "4",
            "Launch",
            "Refresh alerts.",
            "Review second batch company evidence.",
            "Update LinkedIn notes.",
            "Draft second card.",
            "Publish newsletter issue 2.",
            "One new falsifiable judgment is logged.",
        ],
        [
            "5",
            "Weekly rhythm",
            "Alerts, transcript, red flags.",
            "Glassdoor and review themes.",
            "LinkedIn talent-flow notes.",
            "Select 1-2 companies.",
            "Publish bilingual issue.",
            "No more than one main judgment per issue.",
        ],
        [
            "6",
            "Weekly rhythm",
            "Alerts, transcript, red flags.",
            "Glassdoor and review themes.",
            "LinkedIn talent-flow notes.",
            "Select 1-2 companies.",
            "Publish bilingual issue.",
            "At least 10 high-quality external interactions.",
        ],
        [
            "7",
            "Weekly rhythm",
            "Alerts, transcript, red flags.",
            "Glassdoor and review themes.",
            "LinkedIn talent-flow notes.",
            "Select 1-2 companies.",
            "Publish bilingual issue.",
            "Reader feedback log includes 3-5 conversations.",
        ],
        [
            "8",
            "Weekly rhythm",
            "Alerts, transcript, red flags.",
            "Glassdoor and review themes.",
            "LinkedIn talent-flow notes.",
            "Select 1-2 companies.",
            "Publish bilingual issue.",
            "Six-week streak remains intact.",
        ],
        [
            "9",
            "Validation loop",
            "Audit tracker for missing thresholds.",
            "Update baseline coverage map.",
            "Patch weakest company tabs.",
            "Draft recap notes.",
            "Publish bilingual issue.",
            "Every open judgment has one primary metric, threshold, and deadline.",
        ],
        [
            "10",
            "Validation loop",
            "Refresh alerts.",
            "Complete remaining baseline gaps.",
            "Review reader feedback.",
            "Pick follow-up companies.",
            "Publish bilingual issue.",
            "Baseline progress reaches at least 12 companies.",
        ],
        [
            "11",
            "Validation loop",
            "Review tracker statuses.",
            "Prepare methodology recap.",
            "Identify signal/noise lessons.",
            "Draft week-12 review.",
            "Publish bilingual issue.",
            "Do not publish hit-rate percentages before 10 closed judgments.",
        ],
        [
            "12",
            "Review",
            "Close any due judgments.",
            "Update all statuses honestly.",
            "Write method review.",
            "Decide next operating mode.",
            "Publish three-month review.",
            "12-15 baseline files, 8-12 cards, 10 reader feedback points.",
        ],
    ]


def add_table_sheet(workbook: Workbook, title: str, headers: list[str], rows: list[list[object]]) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    style_header(sheet)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    set_widths(sheet)


def style_header(sheet) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def set_widths(sheet) -> None:
    for column_cells in sheet.columns:
        letter = column_cells[0].column_letter
        max_len = max(len(str(cell.value or "")) for cell in column_cells[:100])
        sheet.column_dimensions[letter].width = min(max(max_len + 2, 12), 55)
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def add_validation(sheet, cell_range: str, options: list[str]) -> None:
    formula = '"' + ",".join(options) + '"'
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    sheet.add_data_validation(validation)
    validation.add(cell_range)


def create_readme_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("README", 0)
    rows = [
        ["CultureRadar MVP Workbook", ""],
        ["Purpose", "Single source of truth for the first 12-week manual research run."],
        ["Positioning", "Bilingual public research newsletter for long-term personal investors tracking company culture inflection signals."],
        ["Core Rule", "Do not publish a judgment without one primary validation metric, threshold, deadline, timestamp, and source links."],
        ["Compliance Boundary", "No buy/sell/hold language, target prices, position sizing, personalized advice, or return promises."],
        ["Operating Model", "One research system, English primary output, Chinese summary and commentary."],
        ["Cadence", "Monday alerts/transcripts, Tuesday reviews, Wednesday talent flow, Thursday selection, Friday newsletter."],
        ["Closed Judgment Rule", "Do not publish hit-rate percentages before at least 10 judgments have closed."],
    ]
    for row in rows:
        sheet.append(row)
    sheet["A1"].font = Font(size=16, bold=True, color="1F4E78")
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 110
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def create_card_template_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("Card Template")
    rows = [
        ["Field", "Required Content"],
        ["Company", "Company name"],
        ["Ticker", "Public ticker"],
        ["Direction", "Improving / Stable / Deteriorating"],
        ["Inflection Start", "Month or quarter when the cultural turn appears to begin"],
        ["Confidence", "Low / Medium / High"],
        ["Evidence Line 1", "Talent flow: trend chart note, plain-English interpretation, source link"],
        ["Evidence Line 2", "Employee review trend: rating/sub-score movement and review theme drift"],
        ["Evidence Line 3", "Management language: transcript tone and answer-quality shifts"],
        ["Evidence Line 4", "External red flags: lawsuits, layoffs, SEC, local news, or none material"],
        ["Watch Items", "2-4 observable events to monitor next"],
        ["Primary Validation Metric", "Exactly one metric that determines hit/miss/partial"],
        ["Threshold", "Numeric or objective threshold locked at publish time"],
        ["Deadline", "Date or quarter when the judgment is evaluated"],
        ["Sources", "Clickable links to every material claim"],
        ["Timestamp", "Publish date/time"],
        ["Disclaimer", "Information and investor education only; not investment advice."],
    ]
    for row in rows:
        sheet.append(row)
    style_header(sheet)
    set_widths(sheet)


def create_company_sheet(workbook: Workbook, item: dict[str, object]) -> None:
    sheet = workbook.create_sheet(str(item["ticker"]))
    sheet["A1"] = "Company"
    sheet["B1"] = item["company"]
    sheet["C1"] = "Ticker"
    sheet["D1"] = item["ticker"]
    sheet["E1"] = "Industry"
    sheet["F1"] = item["industry"]
    sheet["A2"] = "Baseline Status"
    sheet["B2"] = "Not started"
    sheet["C2"] = "Pilot Role"
    sheet["D2"] = item["pilot_role"]
    sheet["A4"] = "Current Evidence Baseline"
    sheet["A4"].font = Font(bold=True, color="1F4E78")
    headers = ["Evidence Line", "Current Baseline", "Trend", "Last Updated", "Source URL", "Notes"]
    for col, header in enumerate(headers, start=1):
        sheet.cell(5, col).value = header
    for row_idx, line in enumerate(
        ["Talent Flow", "Employee Review Trend", "Management Language", "External Red Flags"], start=6
    ):
        sheet.cell(row_idx, 1).value = line
    sheet["A11"] = "Quarterly Talent Flow Log"
    sheet["A11"].font = Font(bold=True, color="1F4E78")
    talent_headers = ["Quarter", "Key Inflows", "Key Outflows", "Net Direction", "Roles Covered", "Source Links", "Notes"]
    for col, header in enumerate(talent_headers, start=1):
        sheet.cell(12, col).value = header
    for row in range(13, 21):
        sheet.cell(row, 1).value = ""
    sheet["A23"] = "Employee Review Sampling Log"
    sheet["A23"].font = Font(bold=True, color="1F4E78")
    review_headers = [
        "Date",
        "Overall Rating",
        "CEO Approval",
        "Business Outlook",
        "Management Confidence",
        "Review Themes",
        "Source Link",
    ]
    for col, header in enumerate(review_headers, start=1):
        sheet.cell(24, col).value = header
    sheet["A35"] = "Transcript Reading Log"
    sheet["A35"].font = Font(bold=True, color="1F4E78")
    transcript_headers = ["Call Date", "Quarter", "Growth Language", "Efficiency Language", "Answer Quality", "Source Link", "Notes"]
    for col, header in enumerate(transcript_headers, start=1):
        sheet.cell(36, col).value = header
    sheet["A47"] = "Red Flag / External Signal Log"
    sheet["A47"].font = Font(bold=True, color="1F4E78")
    red_flag_headers = ["Date", "Type", "Materiality", "Description", "Source Link", "Follow-up Needed"]
    for col, header in enumerate(red_flag_headers, start=1):
        sheet.cell(48, col).value = header
    sheet["A58"] = "Draft Judgment"
    sheet["A58"].font = Font(bold=True, color="1F4E78")
    draft_rows = [
        ["Direction", ""],
        ["Inflection Start", ""],
        ["Confidence", ""],
        ["Primary Validation Metric", ""],
        ["Threshold", ""],
        ["Deadline", ""],
        ["Watch Items", ""],
    ]
    for row_offset, row in enumerate(draft_rows, start=59):
        sheet.cell(row_offset, 1).value = row[0]
        sheet.cell(row_offset, 2).value = row[1]
    for header_row in [5, 12, 24, 36, 48]:
        for cell in sheet[header_row]:
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A6"
    set_widths(sheet)
    add_validation(sheet, "C6:C9", ["Improving", "Stable", "Deteriorating", "Mixed", "Unknown"])
    add_validation(sheet, "B59:B59", ["Improving", "Stable", "Deteriorating"])
    add_validation(sheet, "B61:B61", ["Low", "Medium", "High"])


def build_workbook() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)
    create_readme_sheet(workbook)
    add_table_sheet(workbook, "Company Universe", COMPANY_HEADERS, company_rows())
    universe = workbook["Company Universe"]
    add_validation(universe, "E2:E20", ["Yes", "No"])
    add_validation(universe, "F2:F20", ["Core", "Alternate"])
    add_validation(
        universe,
        "M2:M20",
        ["Not started", "In progress", "Baseline complete", "Ready to publish", "Paused"],
    )
    add_table_sheet(workbook, "Judgment Tracker", TRACKER_HEADERS, [])
    tracker = workbook["Judgment Tracker"]
    for _ in range(50):
        tracker.append([""] * len(TRACKER_HEADERS))
    tracker.auto_filter.ref = tracker.dimensions
    add_validation(tracker, "E2:E51", ["Improving", "Stable", "Deteriorating"])
    add_validation(tracker, "F2:F51", ["Low", "Medium", "High"])
    add_validation(tracker, "G2:G51", ["Obvious", "Debatable", "Contrarian"])
    add_validation(tracker, "K2:K51", ["Open", "Hit", "Miss", "Partial", "Void"])
    add_table_sheet(workbook, "Source Library", SOURCE_HEADERS, source_rows())
    add_table_sheet(workbook, "Alert Queries", ALERT_HEADERS, alert_rows())
    add_table_sheet(workbook, "Weekly Cadence", WEEKLY_HEADERS, weekly_rows())
    create_card_template_sheet(workbook)
    for company in selected_companies():
        create_company_sheet(workbook, company)
    workbook.save(DATA_DIR / "CultureRadar_MVP_Workbook.xlsx")


def main() -> None:
    write_csv(DATA_DIR / "company_universe.csv", COMPANY_HEADERS, company_rows())
    write_csv(DATA_DIR / "judgment_tracker.csv", TRACKER_HEADERS, [])
    write_csv(DATA_DIR / "source_library.csv", SOURCE_HEADERS, source_rows())
    write_csv(DATA_DIR / "google_alert_queries.csv", ALERT_HEADERS, alert_rows())
    write_csv(DATA_DIR / "weekly_cadence.csv", WEEKLY_HEADERS, weekly_rows())
    build_workbook()
    print(f"Generated CultureRadar assets in {DATA_DIR}")


if __name__ == "__main__":
    main()
